import openpyxl as xl
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
print(cell.value)
for row in range(1, 2):
    cell = sheet.cell(row, 4)
    new_name = 'New Price'
    new_cell = sheet.cell(row, 4)
    new_cell.value = new_name

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    new_price = cell.value * 0.9
    new_price_cell = sheet.cell(row, 4)
    new_price_cell.value = new_price

wb.save('newtransac.xlsx')
